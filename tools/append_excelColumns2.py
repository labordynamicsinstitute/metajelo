#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 12 10:05:09 2019

@author: flaviostanchi
"""

basepath = '/Users/flaviostanchi/Documents/Workspace/Python/Excel_columns/'

import openpyxl
import sys

def main():
    print("Usage: python append_excelColumns2.py ifile ofile isheet osheet")
    colnames = ['data', 'Data']
    ifile = sys.argv[1]
    ofile = sys.argv[2]
    isheet = sys.argv[3]
    osheet = sys.argv[4]
    
    bookfrom = openpyxl.load_workbook(ifile)
    sheetfrom = bookfrom[isheet]
    bookto = openpyxl.load_workbook(ofile)
    sheetto = bookto[osheet]
    
    n = sheetfrom.max_column  
    m = sheetto.max_column
    r = sheetto.max_row
        
    for i in range(1, n+1):
        if sheetfrom.cell(row=1, column=i).value in colnames:
            for j in range(1, r+1):
                if sheetfrom.cell(row=j, column=1).value == sheetto.cell(row=j, column=1).value:
                    sheetto.cell(row=j, column=m+1).value = sheetfrom.cell(row=j, column=i).value
                else: print("The input spreadsheet doesn't conform to the schema")
            m = sheetto.max_column
    
    bookto.save(basepath + 'tabular-metajelo.xlsx')
    
if __name__ == '__main__':
    main()